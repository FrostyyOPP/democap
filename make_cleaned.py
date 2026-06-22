"""Reproduce the cleaned Brightwave table locally (matches what Copilot did):
standardize dates to DD-MMM-YYYY, remove duplicate rows, flag empty Units cells.
Saves a formatted .xlsx (table style + yellow-flagged gaps) for the reveal shot.
"""
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from dateutil import parser as dparser
from datetime import datetime

SRC = r"C:\Users\starw\Downloads\OneDrive_1_6-22-2026\Brightwave_Sales_RAW.xlsx"
OUT = r"C:\Users\starw\Downloads\Claude\democap\runs\Brightwave_Sales_CLEANED.xlsx"

wb = openpyxl.load_workbook(SRC)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
data = rows[1:]


def fmt_date(d):
    if isinstance(d, datetime):
        return d.strftime("%d-%b-%Y")
    try:
        return dparser.parse(str(d), dayfirst=True).strftime("%d-%b-%Y")
    except Exception:
        return d


seen, clean = set(), []
for r in data:
    if all(c is None for c in r):
        continue
    r = list(r)
    r[0] = fmt_date(r[0])
    if isinstance(r[3], str):
        r[3] = r[3].strip().title()      # Product: plus -> Plus
    if isinstance(r[1], str):
        r[1] = r[1].strip()
    key = tuple(r)
    if key in seen:
        continue
    seen.add(key)
    clean.append(r)

print(f"rows in: {len(data)}  ->  clean rows out: {len(clean)}  (removed {len(data)-len(clean)} dupes/blanks)")

out = openpyxl.Workbook()
o = out.active
o.title = "Sales"
o.append(header)
yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
empty_units = 0
for i, r in enumerate(clean, start=2):
    o.append(r)
    if r[4] is None or str(r[4]).strip() == "":
        o.cell(row=i, column=5).fill = yellow   # flag empty Units
        empty_units += 1
print(f"flagged {empty_units} empty Units cells")

# widths + table style
widths = {"A": 14, "B": 12, "C": 16, "D": 13, "E": 8, "F": 12, "G": 12}
for col, w in widths.items():
    o.column_dimensions[col].width = w
ref = f"A1:G{len(clean)+1}"
tbl = Table(displayName="CleanTable", ref=ref)
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
o.add_table(tbl)
out.save(OUT)
print("saved:", OUT)
